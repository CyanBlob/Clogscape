#!/usr/bin/env python3
"""
Parse the Bountyscape crowdsourcing spreadsheet and emit JSON.

Dependencies: openpyxl (pip install openpyxl)

Enhancements:
- Parses shorthand numbers like "10k" -> 10000, "2.5m" -> 2500000, "1b" -> 1000000000.
"""

import argparse
import json
import math
import re
from typing import Any, Dict, Optional

from openpyxl import load_workbook


DIFFICULTY_ENUM = {
    "Novice": 0,
    "Easy": 1,
    "Medium": 2,
    "Hard": 3,
    "Expert": 4,
    "Grandmaster": 5,
}

BOUNTY_TYPE_ENUM = {
    "Boss": 0,
    "Raid": 1,
    "Clue": 2,
    "Minigame": 3,
    "Grind": 4,
    "Skilling": 5,
    "Challenge": 6,
    "Misc": 7,
    "Fetch": 8,
    "RealLife": 9,
    "Slayer": 10,
}

EXPECTED_HEADERS = [
    "Name",
    "Description",
    "Difficulty",
    "Bounty Type",
    "Count Min",
    "Count Max",
    "Key Reward Min",
    "Key Reward Max",
    "Extra Key Chance (%)",
    "Max Lifetime Keys",
    "GP Reward Min",
    "GP Reward Max",
]

# e.g. "10k", "2.5m", "1b", with optional commas/spaces
SHORTHAND_RE = re.compile(r"^\s*([0-9]+(?:\.[0-9]+)?)\s*([kKmMbB])\s*$")


def is_blank(x: Any) -> bool:
    return x is None or (isinstance(x, str) and x.strip() == "")


def parse_number_like(x: Any) -> Optional[float]:
    """
    Returns a float for numeric-like inputs, including:
      - int/float cells
      - "10k" / "2.5m" / "1b"
      - strings like "10,000" or "  42 "
    Returns None for blanks.
    """
    if is_blank(x):
        return None

    if isinstance(x, bool):
        return float(int(x))

    if isinstance(x, (int, float)):
        if isinstance(x, float) and math.isnan(x):
            return None
        return float(x)

    s = str(x).strip()
    if s == "":
        return None

    m = SHORTHAND_RE.match(s)
    if m:
        base = float(m.group(1))
        suffix = m.group(2).lower()
        mult = {"k": 1e3, "m": 1e6, "b": 1e9}[suffix]
        return base * mult

    # remove commas/underscores commonly used as separators
    s2 = s.replace(",", "").replace("_", "")
    return float(s2)


def to_int_or_none(x: Any) -> Optional[int]:
    n = parse_number_like(x)
    if n is None:
        return None
    return int(round(n))


def to_float_or_none(x: Any) -> Optional[float]:
    return parse_number_like(x)


def find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 50) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = {str(v).strip(): idx for idx, v in enumerate(values) if not is_blank(v)}
        if "Name" in normalized and "Difficulty" in normalized and "Bounty Type" in normalized:
            return r
    raise RuntimeError("Could not locate header row (looked in first 50 rows).")


def build_column_map(ws, header_row: int) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if is_blank(v):
            continue
        header = str(v).strip()
        if header in EXPECTED_HEADERS:
            col_map[header] = c
    missing = [h for h in EXPECTED_HEADERS if h not in col_map]
    if missing:
        raise RuntimeError(f"Missing expected columns in header row: {missing}")
    return col_map


def parse_row(ws, r: int, col: Dict[str, int]) -> Optional[Dict[str, Any]]:
    name = ws.cell(row=r, column=col["Name"]).value
    if is_blank(name):
        return None

    desc = ws.cell(row=r, column=col["Description"]).value
    difficulty = ws.cell(row=r, column=col["Difficulty"]).value
    bounty_type = ws.cell(row=r, column=col["Bounty Type"]).value

    count_min = ws.cell(row=r, column=col["Count Min"]).value
    count_max = ws.cell(row=r, column=col["Count Max"]).value
    min_keys = ws.cell(row=r, column=col["Key Reward Min"]).value
    max_keys = ws.cell(row=r, column=col["Key Reward Max"]).value
    extra_key_pct = ws.cell(row=r, column=col["Extra Key Chance (%)"]).value
    max_lifetime_keys = ws.cell(row=r, column=col["Max Lifetime Keys"]).value
    min_gp = ws.cell(row=r, column=col["GP Reward Min"]).value
    max_gp = ws.cell(row=r, column=col["GP Reward Max"]).value

    diff_str = "" if is_blank(difficulty) else str(difficulty).strip()
    bt_str = "" if is_blank(bounty_type) else str(bounty_type).strip()

    if diff_str not in DIFFICULTY_ENUM:
        raise RuntimeError(f"Row {r}: unknown Difficulty '{diff_str}'")
    if bt_str not in BOUNTY_TYPE_ENUM:
        raise RuntimeError(f"Row {r}: unknown Bounty Type '{bt_str}'")

    obj: Dict[str, Any] = {
        # From sheet (unchanged text)
        "name": str(name),
        "description": None if is_blank(desc) else str(desc),
        "difficulty": DIFFICULTY_ENUM[diff_str],
        "bountyType": BOUNTY_TYPE_ENUM[bt_str],
        "minKeys": to_int_or_none(min_keys) or 0,
        "maxKeys": to_int_or_none(max_keys) or 0,
        "extraKeyChance": to_float_or_none(extra_key_pct) or 0.0,
        "maxLifetimeKeys": to_int_or_none(max_lifetime_keys),
        "minGp": to_int_or_none(min_gp) or 0,
        "maxGp": to_int_or_none(max_gp) or 0,

        # Optional count fields (from sheet)
        "countMin": to_int_or_none(count_min),
        "countMax": to_int_or_none(count_max),

        # Missing-from-sheet fields (defaults)
        "keyChance": 0.0,
        "skipChance": 0.0,
        "lastCompleted": None,
        "lastRolled": None,
        "combatLevel": 3,
        "isWildy": False,
        "requirementLocked": False,
        "completedLocked": False,
    }

    return obj


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        nargs="?",
        default="/mnt/data/Bountyscape bounties crowdsourcing.xlsx",
        help="Path to the .xlsx file",
    )
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    ap.add_argument("-o", "--out", default="bounties.json", help="Output JSON filename")
    ap.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    header_row = find_header_row(ws)
    col_map = build_column_map(ws, header_row)

    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_obj = parse_row(ws, r, col_map)
        if row_obj is not None:
            out.append(row_obj)

    with open(args.out, "w", encoding="utf-8") as f:
        if args.pretty:
            json.dump(out, f, ensure_ascii=False, indent=2)
        else:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    print(f"Wrote {len(out)} tasks -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
